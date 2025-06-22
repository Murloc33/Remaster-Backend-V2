import shutil
from datetime import datetime
from sqlite3 import Connection

from fastapi import APIRouter, Depends, Body, Path
from openpyxl import load_workbook
from starlette.responses import Response
from typing_extensions import Annotated

from core.methods import get_connection, resource_path
from modules.databases.methods import convert_date
from modules.databases.schemes import Databases

router = APIRouter(prefix='/databases')


@router.get('/')
def read_database(connection: Annotated[Connection, Depends(get_connection)]):
    cursor = connection.cursor()

    cursor.execute("Select * from databases")
    result = cursor.fetchall()

    return {"data": Databases.validate_python(result)}


@router.put('/doping-athletes/upload')
def upload_doping_athlete(
    path: Annotated[str, Body(embed=True)],
    connection: Annotated[Connection, Depends(get_connection)]
):
    cursor = connection.cursor()

    cursor.execute('DELETE FROM doping_athletes')
    cursor.execute('UPDATE sqlite_sequence SET seq = 0 WHERE name = ?', ('doping_athletes',))
    connection.commit()

    wb = load_workbook(path)

    for row in wb["Лист1"].iter_rows(min_row=4, values_only=True):
        if not any(row):
            break

        cursor.execute(
            "INSERT INTO  doping_athletes "
            "(full_name, sport, birth_date, violation_description, disqualification_duration, disqualification_start, disqualification_end)"
            " VALUES (?, ?, ?, ?, ?, ?, ?)",
            (
                row[0], row[2].capitalize(), convert_date(row[1]), row[3], row[4].capitalize(),
                convert_date(row[5]), convert_date(row[6])
            )
        )

    shutil.copy(path, resource_path("resources/doping-athletes.xlsx"))

    file_name = path.rsplit('\\', 1)[1]

    cursor = connection.cursor()
    cursor.execute(
        "UPDATE databases SET date = ?, file_name = ? WHERE slug = ?",
        (str(datetime.now().strftime('%d.%m.%Y')), file_name, "doping-athletes")
    )
    connection.commit()

    return Response()


@router.put('/{slug}/upload')
def upload_sports(
    slug: Annotated[str, Path()],
    path: Annotated[str, Body(embed=True)],
    connection: Annotated[Connection, Depends(get_connection)]
):
    file_name = path.rsplit('\\', 1)[1]
    extension = file_name.rsplit('.', 1)[1]

    shutil.copy(path, resource_path(f'resources/{slug}.{extension}'))

    cursor = connection.cursor()
    cursor.execute(
        "UPDATE databases SET date = ?, file_name = ? WHERE slug = ?",
        (str(datetime.now().strftime('%d.%m.%Y')), file_name, slug)
    )
    connection.commit()

    return Response()


@router.put('/{slug}/download')
def download_file(
    slug: Annotated[str, Path()],
    path: Annotated[str, Body(embed=True)],
):
    file_name = path.rsplit('\\', 1)[1]
    extension = file_name.rsplit('.', 1)[1]

    shutil.copy(resource_path(f"resources/{slug}.{extension}"), path)
    return Response()
