import sqlite3
from pathlib import Path
from typing import Dict

from openpyxl import load_workbook


def get_or_create_sport(cursor, sport_name: str) -> int:
    """
    Get sport ID by name, or create a new sport if it doesn't exist.
    Returns (sport_id, is_new)
    """
    # Try to find existing sport
    cursor.execute("SELECT id FROM sports WHERE LOWER(name) = LOWER(?)", (sport_name,))
    result = cursor.fetchone()

    if result:
        return result[0], False

    # Create new sport
    cursor.execute("INSERT INTO sports (name) VALUES (?)", (sport_name,))
    return cursor.lastrowid, True


def get_sport_mapping(cursor) -> Dict[str, int]:
    """Get mapping of lowercase sport names to their IDs from the database."""
    cursor.execute("SELECT id, name FROM sports")
    return {name.lower(): id for id, name in cursor.fetchall()}


def extract_sport_name(org_name: str) -> str:
    """
    Extract sport name from organization name.
    Returns the extracted sport name in title case.
    """
    # Common prefixes/suffixes to remove
    prefixes = [
        "Общероссийская общественная организация «Федерация",
        "Общероссийская спортивная федерация",
        "Федерация",
        "Всероссийская федерация",
        "Общероссийская федерация",
        "ФЕДЕРАЦИЯ",
        "федерация",
    ]

    # Remove common prefixes
    name = org_name.strip()
    for prefix in prefixes:
        if name.lower().startswith(prefix.lower()):
            name = name[len(prefix):].strip("» ", )

    # Remove anything in parentheses or after запятая/дефис
    name = name.split("(")[0].split(",")[0].split(" - ")[0].strip()

    # Capitalize first letter of each word
    return ' '.join(word.capitalize() for word in name.split())


def import_organizations(excel_path: str, db_path: str):
    print(f"Reading Excel file: {excel_path}")

    # Connect to the database
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # Get existing sport mappings
    sport_mapping = get_sport_mapping(cursor)
    print(f"Found {len(sport_mapping)} existing sports in database")

    try:
        # Load the workbook and get the active worksheet
        wb = load_workbook(filename=excel_path, read_only=True)
        ws = wb.active

        # Get header row to find the correct column index
        headers = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        print("Columns in Excel file:", headers)

        # Find the index of the organization name column
        org_col_index = 4

        # Prepare data for insertion
        organizations_to_add = []
        added_count = 0
        skipped_count = 0
        new_sports_created = 0

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                org_name = str(row[org_col_index - 1]).strip()

                # Skip empty rows
                if not org_name:
                    print(f"Skipping row {row_num}: Empty organization name")
                    skipped_count += 1
                    continue

                # Extract and normalize sport name
                sport_name = str(row[1]).strip()

                if not sport_name:
                    print(f"Skipping row {row_num}: Could not extract sport name from: {org_name}")
                    skipped_count += 1
                    continue

                # Get or create sport
                sport_id, is_new = get_or_create_sport(cursor, sport_name)

                if is_new:
                    print(f"Created new sport: {sport_name} (ID: {sport_id})")
                    sport_mapping[sport_name.lower()] = sport_id
                    new_sports_created += 1

                # Add organization to the batch
                organizations_to_add.append((org_name, sport_id))
                added_count += 1
            except Exception as e:
                print(f"Error processing row {row_num}: {e}")
                skipped_count += 1
                continue

        # Insert any remaining organizations
        if organizations_to_add:
            cursor.executemany(
                "INSERT INTO organizations (title, sport_id) VALUES (?, ?)",
                organizations_to_add
            )

        conn.commit()
        print(f"\nImport summary:")
        print(f"- Successfully added: {added_count} organizations")
        print(f"- New sports created: {new_sports_created}")
        print(f"- Skipped rows: {skipped_count}")

    except Exception as e:
        print(f"Error during processing: {e}")
        conn.rollback()
        raise

    finally:
        # Close the workbook and database connection
        if 'wb' in locals():
            wb.close()
        conn.close()


if __name__ == "__main__":
    # Paths - adjust as needed
    base_dir = Path(__file__).parent.parent
    excel_file = base_dir / "resources" / "1. РЕЕСТР федераций май 2025.xlsx"
    db_file = base_dir / "resources" / "remaster.db"

    print(f"Starting import from {excel_file}")
    import_organizations(str(excel_file), str(db_file))
    print("Import completed.")
